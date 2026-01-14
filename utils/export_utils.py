"""
Export Utilities
Handles data export to Excel and other formats
"""

from typing import List, Dict
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ExcelExporter:
    """Export RFID read history to Excel format"""
    
    # Styling - only initialized if openpyxl is available
    HEADER_FILL = None
    HEADER_FONT = None
    BORDER = None
    
    @staticmethod
    def is_available() -> bool:
        """Check if export functionality is available"""
        return HAS_OPENPYXL
    
    @staticmethod
    def export_read_history(data: List[Dict], filepath: str) -> tuple[bool, str]:
        """
        Export read history to Excel file
        
        Args:
            data: List of dicts with keys: index, antenna, epc, rssi, timestamp, s1, s2
            filepath: Output file path
            
        Returns:
            (success, message) tuple
        """
        if not HAS_OPENPYXL:
            return False, "openpyxl not installed. Run: pip install openpyxl"
        
        if not data:
            return False, "No data to export"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Read History"
            
            # Headers
            headers = ["STT", "Antenna", "EPC", "RSSI", "Timestamp", "S1", "S2"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = ExcelExporter.HEADER_FILL
                cell.font = ExcelExporter.HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
                cell.border = ExcelExporter.BORDER
            
            # Data
            for row_idx, item in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=item['index']).border = ExcelExporter.BORDER
                ws.cell(row=row_idx, column=2, value=item['antenna']).border = ExcelExporter.BORDER
                ws.cell(row=row_idx, column=3, value=item['epc']).border = ExcelExporter.BORDER
                ws.cell(row=row_idx, column=4, value=item['rssi']).border = ExcelExporter.BORDER
                ws.cell(row=row_idx, column=5, value=item['timestamp']).border = ExcelExporter.BORDER
                ws.cell(row=row_idx, column=6, value=str(item['s1'])).border = ExcelExporter.BORDER
                ws.cell(row=row_idx, column=7, value=str(item['s2'])).border = ExcelExporter.BORDER
            
            # Adjust column widths
            column_widths = [8, 10, 30, 8, 15, 8, 8]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Save
            wb.save(filepath)
            return True, f"Exported successfully to {filepath}"
            
        except Exception as e:
            return False, f"Export failed: {str(e)}"
    
    @staticmethod
    def export_detected_tags(tags: List[Dict], filepath: str) -> tuple[bool, str]:
        """
        Export detected tags with confidence scores
        
        Args:
            tags: List of dicts with keys: epc, confidence_ant1, confidence_ant2, confidence_all, direction
            filepath: Output file path
        """
        if not HAS_OPENPYXL:
            return False, "openpyxl not installed"
        
        if not tags:
            return False, "No tags to export"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Detected Tags"
            
            # Headers
            headers = ["EPC", "REL1", "REL2", "REL&", "Direction"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = ExcelExporter.HEADER_FILL
                cell.font = ExcelExporter.HEADER_FONT
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, tag in enumerate(tags, 2):
                ws.cell(row=row_idx, column=1, value=tag['epc'])
                ws.cell(row=row_idx, column=2, value=f"{tag['confidence_ant1']:.1f}")
                ws.cell(row=row_idx, column=3, value=f"{tag['confidence_ant2']:.1f}")
                ws.cell(row=row_idx, column=4, value=f"{tag['confidence_all']:.1f}")
                ws.cell(row=row_idx, column=5, value=tag['direction'])
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 30
            for col in ['B', 'C', 'D', 'E']:
                ws.column_dimensions[col].width = 12
            
            wb.save(filepath)
            return True, f"Exported successfully to {filepath}"
            
        except Exception as e:
            return False, f"Export failed: {str(e)}"
    
    @staticmethod
    def generate_filename(prefix: str = "rfid_export") -> str:
        """Generate a timestamped filename"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.xlsx"


# Initialize styling if openpyxl is available
if HAS_OPENPYXL:
    ExcelExporter.HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ExcelExporter.HEADER_FONT = Font(bold=True, color="FFFFFF")
    ExcelExporter.BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
